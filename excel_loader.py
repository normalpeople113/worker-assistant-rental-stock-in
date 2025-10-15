from openpyxl import load_workbook

def load_rows_with_index(path):
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        imei, status, location = row[:3]
        rows.append({
            "IMEI": imei,
            "Status": status,
            "Location": location,
            "row_index": i
        })
    return wb, ws, rows
